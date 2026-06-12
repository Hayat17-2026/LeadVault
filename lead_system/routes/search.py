from services.auth_service import login_required, rate_limited
from flask import Blueprint, render_template, request, jsonify
from services.search_service import run_search
from services.activity_service import log_activity
from database.db import get_db

search_bp = Blueprint("search", __name__)

@search_bp.route("/", methods=["GET"])
@login_required
@rate_limited
def search_page():
    return render_template("search.html")

@search_bp.route("/run", methods=["POST"])
@login_required
@rate_limited
def run():
    """
    Step 1 — Research Module
    Receives: keywords, platform, region
    Returns:  list of candidate leads found from public data
    """
    try:
        data     = request.get_json(force=True) or {}
        keywords = data.get("keywords", [])
        platform = data.get("platform", "web")
        region   = data.get("region", "global")

        results = run_search(keywords, platform, region)

        db = get_db()
        db.execute(
            "INSERT INTO search_logs (query, platform, results) VALUES (?, ?, ?)",
            (", ".join(keywords), platform, len(results))
        )
        db.commit()
        db.close()

        log_activity("search", f'Searched "{", ".join(keywords)}" — {len(results)} results')
        return jsonify({"status": "ok", "results": results, "count": len(results)})
    except Exception as e:
        import traceback
        print(f"[search/run] EXCEPTION: {e}\n{traceback.format_exc()}")
        return jsonify({"status": "error", "message": str(e)}), 500


@search_bp.route("/verify-email", methods=["POST"])
@login_required
def verify_email():
    """Check if an email's domain actually exists (DNS lookup). No external API needed."""
    import re, socket
    data  = request.get_json(force=True) or {}
    email = data.get("email", "").strip().lower()

    fmt = re.compile(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z]{2,}$")
    if not email or not fmt.match(email):
        return jsonify({"status": "ok", "result": "invalid", "reason": "Invalid email format"})

    domain = email.split("@")[1]
    # Common throwaway / known-bad domains
    disposable = {"mailinator.com","guerrillamail.com","trashmail.com","temp-mail.org","yopmail.com","fakeinbox.com"}
    if domain in disposable:
        return jsonify({"status": "ok", "result": "risky", "reason": "Disposable email domain"})

    try:
        socket.setdefaulttimeout(4)
        socket.gethostbyname(domain)
        return jsonify({"status": "ok", "result": "valid", "reason": f"Domain {domain} resolves"})
    except socket.error:
        return jsonify({"status": "ok", "result": "invalid", "reason": f"Domain {domain} does not exist"})


@search_bp.route("/lead-summary", methods=["POST"])
@login_required
@rate_limited
def lead_summary():
    """Generate an AI summary for a single lead using Claude (falls back to rule-based)."""
    import os
    data = request.get_json(force=True) or {}
    lead = data.get("lead", {})

    name      = lead.get("name", "Unknown")
    platform  = lead.get("platform", "web")
    email     = lead.get("email") or "not available"
    phone     = lead.get("phone") or "not available"
    interests = lead.get("interests") or "general"
    score     = lead.get("score", 0)
    snippet   = lead.get("snippet") or ""
    url       = lead.get("source_url") or ""

    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if api_key:
        try:
            import anthropic
            client = anthropic.Anthropic(api_key=api_key)
            prompt = (
                f"Analyze this business lead in 2-3 sentences and give a specific recommended action:\n\n"
                f"Name: {name}\nPlatform: {platform}\nEmail: {email}\nPhone: {phone}\n"
                f"Interests: {interests}\nScore: {score}/100\nSnippet: {snippet}\n\n"
                f"End with: <strong>Recommended action:</strong> [specific action]"
            )
            resp = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=180,
                messages=[{"role": "user", "content": prompt}]
            )
            return jsonify({"status": "ok", "summary": resp.content[0].text})
        except Exception as e:
            pass  # fall through to rule-based

    # Rule-based fallback
    tier    = "High-priority" if score >= 75 else "Medium-priority" if score >= 50 else "Low-priority"
    contact = []
    if lead.get("email"): contact.append("email")
    if lead.get("phone"): contact.append("phone")
    contact_str = " and ".join(contact) if contact else "no contact info"
    action  = "Send a cold email immediately." if lead.get("email") else ("Call directly." if lead.get("phone") else "Find contact details first.")
    summary = (f"<strong>{name}</strong> is a <strong>{tier}</strong> lead found on {platform} "
               f"with {contact_str}. Interests: {interests}. Score: {score}/100.<br>"
               f"<strong>Recommended action:</strong> {action}")
    return jsonify({"status": "ok", "summary": summary})


@search_bp.route("/export-excel", methods=["POST"])
@login_required
@rate_limited
def export_search_excel():
    """Export current search results to a real .xlsx Excel file."""
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    data    = request.get_json(force=True)
    results = data.get("results", [])
    fields  = data.get("fields", "all")
    if not results:
        return jsonify({"status": "error", "message": "No results to export"}), 400

    # (header_label, result_key)  — None key means row index
    COLS = {
        "all":        [("#",None),("Name","name"),("Email","email"),("Phone","phone"),("Platform","platform"),("Interests","interests"),("Source URL","source_url")],
        "name":       [("#",None),("Name","name")],
        "email":      [("#",None),("Email","email")],
        "phone":      [("#",None),("Phone","phone")],
        "name_email": [("#",None),("Name","name"),("Email","email")],
        "name_phone": [("#",None),("Name","name"),("Phone","phone")],
    }
    cols = COLS.get(fields, COLS["all"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Search Results"

    # Header row
    ws.append([c[0] for c in cols])
    purple = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    for col_idx in range(1, len(cols) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = purple
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for i, r in enumerate(results, 1):
        ws.append([i if key is None else r.get(key, "") for _, key in cols])

    # Column widths
    WIDTH_MAP = {"#":5,"Name":30,"Email":28,"Phone":18,"Platform":14,"Interests":22,"Source URL":40}
    for col_idx, (hdr, _) in enumerate(cols, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = WIDTH_MAP.get(hdr, 20)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    FILENAMES = {
        "all":"All_Fields","name":"Names_Only","email":"Emails_Only",
        "phone":"Phones_Only","name_email":"Name_Email","name_phone":"Name_Phone"
    }
    filename = f"LeadVault_{FILENAMES.get(fields,'Export')}.xlsx"
    log_activity("export", f"Exported {len(results)} results ({fields}) to Excel")
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)